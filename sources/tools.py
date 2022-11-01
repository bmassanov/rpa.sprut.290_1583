import ctypes
import shutil
from pathlib import Path
from time import sleep
from typing import Iterable

import xlrd
from openpyxl import load_workbook, Workbook

from settings.paths import app_path
from settings.settings import app_login, app_pass, app_base
from sources.rpamini import App, Clipboard

D_TIMEOUT = 50.0
current_cursor = 0

if ctypes.windll.user32.GetKeyboardLayout(0) != 67699721:
    raise Exception('Смените раскладку на ENG')


class Sprut(App):
    SEARCH = '^F'
    FILTER = '^%{F11}'
    FILTER_RESET = '^%{F12}'

    CLEAR = '{END}+{HOME}{DELETE}'
    COPY_ROWS = '+%{INSERT}'
    COPY_VALUE = '^{INSERT}'
    EDIT = '{F4}'
    SHIFT_EDIT = '+{F4}'
    DUPLICATE = '^{F2}'

    PANE_TOP = 1
    PANE_ATTR_LEFT = 7
    PANE_ATTR_RIGHT = 6
    PANE_ADD = 4
    PANE_DOST_LEFT = 6
    PANE_DOST_RIGHT = 12

    Clipboard = Clipboard

    def __init__(self):
        self.app_path = app_path
        self.app_login = app_login
        self.app_pass = app_pass
        self.app = self.execute()

        # self.serialisation_path = serialisation_path.joinpath(f'{environ.get("iteration")}.json')
        self.serialisation = {}
        self.title = None

    # ? START SPRUT

    def quit(self):
        self.kill_exe('EXCEL.EXE')
        self.kill_exe('sprut.exe')
        self.kill_exe('ref_.*.exe')

    def execute(self):
        self.quit()
        app = self.start_exe(self.app_path)
        return app

    def authorize(self):
        selector_ = [{"title": "Регистрация", "index": 0},
                     {"title": "", "control_type": "Edit", "index": 0}]
        self.find_element(selector_).type_keys(self.format_str(self.app_login))

        selector_ = [{"title": "Регистрация", "index": 0},
                     {"title": "", "control_type": "Edit", "index": 1}]
        self.find_element(selector_).type_keys(self.format_str(self.app_pass))

        base = app_base
        selector_ = [{"title": "Регистрация", "index": 0},
                     {"control_type": "Pane", "index": 5}]
        self.find_element(selector_).type_keys(base)
        selector_ = [{"title": "", "index": 0},
                     {"title": "", "control_type": "ListItem", "index": 0}]
        self.find_element(selector_).type_keys(base + self.Keys.ENTER, set_foreground=False)

        selector_ = [{"title": "Регистрация", "index": 0},
                     {"title": "Ввод", "control_type": "Button", "index": 0}]
        self.find_element(selector_).click_input()

        selector_ = [{"title": "\"Главное меню ПС СПРУТ\"", "index": 0}]
        self.find_element(selector_, timeout=90).maximize()

        return self

    def open_module(self, title: str = 'Шаблоны и правила'):
        self.title = title

        selector_ = [{"title": "\"Главное меню ПС СПРУТ\"", "index": 0}]
        self.find_element(selector_).click_input()
        self.find_element(selector_).type_keys(self.SEARCH)
        if not self.wait_element([{"title": "Поиск", "index": 0}]):
            raise Exception('Окно поиска не открылось.')

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "", "control_type": "Edit", "index": 0}]
        self.find_element(selector_).type_keys(self.format_str(self.title, True), with_spaces=True)

        selector_ = [{"title": "Поиск", "index": 0},
                     {"control_type": "Pane", "index": 9}]
        self.find_element(selector_).type_keys(self.COPY_ROWS)

        data_ = self.clipboard_get(row=True)
        if data_.get('Название модуля') != self.title:
            raise Exception(f'Модуль "{self.title}" не найден')

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "Перейти", "control_type": "Button", "index": 0}]
        self.find_element(selector_).click_input()

        selector_ = [{"title": "Поиск", "index": 0}]
        self.wait_element(selector_, appear=False)

        selector_ = [{"title": "\"Главное меню ПС СПРУТ\"", "index": 0}]
        self.find_element(selector_).type_keys(self.Keys.ENTER)

        selector_ = [{"title": f"\"{self.title}\"", "index": 0}]
        self.wait_element(selector_)

        return self

    # ? SERIALISATION

    # def serialisation_write(self, key, value):
    #     if self.serialisation_path.is_file():
    #         self.serialisation = Json.read(self.serialisation_path.__str__())
    #     self.serialisation[key] = value
    #     Json.write(self.serialisation_path.__str__(), self.serialisation)
    #
    # def serialisation_read(self):
    #     if self.serialisation_path.is_file():
    #         self.serialisation = Json.read(self.serialisation_path.__str__())
    #     return self.serialisation

    # ? SYSTEM

    def close(self):
        selector_ = [{"title": "Редактировать запись", "index": 0}]
        if self.wait_element(selector_, timeout=0):
            self.find_element(selector_).close()

        selector_ = [{"title": f"\"{self.title}\"", "index": 0}]
        self.find_element(selector_).close()

    def clipboard_get(self, row=False, delay_before=0.5, value_error=''):
        sleep(delay_before)
        if not row:
            try:
                result = self.Clipboard.get()
            except (Exception,):
                result = None
            if result is None:
                raise ValueError(value_error)
        else:
            result = {row_.split('\t')[0]: row_.split('\t')[1] for row_ in self.Clipboard.get().split('\r\n')}
        self.Clipboard.set('')
        return result

    def format_str(self, value: str, enter=False, clear=True, backspace=True):
        value = self.protect_str(str(value))
        if backspace:
            value = self.Keys.BACKSPACE * 2 + value
        if clear:
            value = self.CLEAR + value
        if enter:
            value = value + self.Keys.ENTER
        return value

    def get_pane(self, index):
        selector_ = [{"title": f"\"{self.title}\"", "index": 0},
                     {"title": "", "control_type": "Pane", "index": index}]
        return self.find_element(selector_)

    # ? EDIT CELLS

    def get_input(self, index):
        selector_ = [{"title": "Редактировать запись", "index": 0},
                     {"control_type": "Edit", "index": index}]
        element_ = self.find_element(selector_)
        value_ = element_.iface_value.CurrentValue
        return value_

    def set_input(self, index, value):
        selector_ = [{"class_name": "TfrmParams", "index": 0},
                     {"control_type": "Edit", "index": index}]
        element_ = self.find_element(selector_)
        value_ = element_.iface_value.CurrentValue

        element_.click_input()
        element_.type_keys(self.Keys.BACKSPACE * len(value_) + self.Keys.DELETE * len(value_))
        element_.type_keys(self.format_str(value), with_spaces=True)

        selector_ = [{"class_name": "TfrmParams", "index": 0},
                     {"title": "", "control_type": "TitleBar", "index": 0}]
        self.find_element(selector_).click_input()

        return value_, value

    def set_select(self, index, value):
        original_selector_ = [{"class_name": "TfrmParams", "index": 0},
                              {"control_type": "Edit", "index": index}]
        original_value = self.find_element(original_selector_).iface_value.CurrentValue
        self.find_element(original_selector_).click_input()
        self.find_element(original_selector_).type_keys(self.SEARCH)
        if not self.wait_element([{"title": "Поиск", "index": 0}]):
            self.find_element(original_selector_).type_keys(self.SEARCH)
        if not self.wait_element([{"title": "Поиск", "index": 0}]):
            raise Exception('Окно поиска не открылось.')

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "", "control_type": "Edit", "index": 0}]
        value_ = self.format_str(value, enter=True, clear=False)
        self.find_element(selector_).click_input()
        self.find_element(selector_).type_keys(value_, with_spaces=True, set_foreground=True)

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "", "control_type": "Pane", "index": 4}]
        self.find_element(selector_).type_keys(self.COPY_VALUE)
        self.clipboard_get(row=False, value_error=f'не найдено зн.: {value}')

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "Перейти", "control_type": "Button", "index": 0}]
        self.find_element(selector_).click_input()

        timeout_ = float(10)
        while self.find_element(original_selector_).iface_value.CurrentValue != value and timeout_ > 0:
            sleep(1)
            timeout_ -= 1

        selector_ = [{"class_name": "TfrmParams", "index": 0},
                     {"title": "", "control_type": "TitleBar", "index": 0}]
        self.find_element(selector_).click_input()

        return original_value, value

    def set_multiselect(self, index, params: Iterable):
        original_selector_ = [{"class_name": "TfrmParams", "index": 0},
                              {"control_type": "Edit", "index": index}]
        original_value = self.find_element(original_selector_).iface_value.CurrentValue
        self.find_element(original_selector_).click_input()
        element_ = self.find_element(original_selector_)
        try:
            element_.type_keys(self.SEARCH)
        except (Exception,):
            sleep(10)
        finally:
            self.wait_cursor(10)

        if not self.wait_element([{"title": "Поиск", "index": 0}]):
            self.find_element(original_selector_).type_keys(self.SEARCH)
        if not self.wait_element([{"title": "Поиск", "index": 0}]):
            raise Exception('Окно поиска не открылось.')

        selector_ = [{"class_name": "Tvms_search_fm_builder", "index": 0},
                     {"title": "", "control_type": "Button", "index": 3}]
        self.find_element(selector_).click_input()

        for param in params:
            selector_ = [{"class_name": "Tvms_search_fm_builder", "index": 0},
                         {"control_type": "Pane", "index": 8}]
            self.find_element(selector_).click_input()

            category, down, value, click = param[0], param[1], param[2], param[3]

            selector_ = [{"class_name": "TcxComboBoxPopupWindow", "index": 0},
                         {"title": "", "control_type": "ListItem", "index": category}]
            if down:
                self.find_element(selector_).type_keys(self.keys.PAGE_DOWN * down, set_foreground=False)
            sleep(1.5)
            self.find_element(selector_).click_input()

            selector_ = [{"title": "Поиск", "index": 0},
                         {"title": "", "control_type": "Edit", "index": 0}]
            value_ = self.format_str(value, enter=True, clear=True)
            self.find_element(selector_).click_input()
            self.find_element(selector_).type_keys(value_, with_spaces=True, set_foreground=True)

            selector_ = [{"title": "Поиск", "index": 0},
                         {"title": "", "control_type": "Pane", "index": 4}]
            self.find_element(selector_).type_keys(self.COPY_VALUE)
            self.clipboard_get(row=False, value_error=f'не найдено зн.:')

            selector_ = [{"class_name": "Tvms_search_fm_builder", "index": 0},
                         {"title": "", "control_type": "Button", "index": 1}]
            if click:
                self.find_element(selector_).click_input()

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "Выделить все", "control_type": "Button", "index": 0}]
        self.find_element(selector_).click_input()

        selector_ = [{"class_name": "TfrmParams", "index": 0},
                     {"title": "", "control_type": "TitleBar", "index": 0}]
        self.find_element(selector_).click_input()

        return original_value

    def f4(self, long_name=None, short_name=None, condition=None, category=None, sign=None, area=None):
        selector_ = [{"title": f"\"{self.title}\"", "index": 0}]
        self.find_element(selector_).type_keys(self.EDIT)
        result = {}

        if long_name is not None:
            result['long_name'] = dict()
            result['long_name']['original'], result['long_name']['new'] = self.set_input(5, long_name)

        if short_name is not None:
            result['short_name'] = dict()
            result['short_name']['original'], result['short_name']['new'] = self.set_input(4, short_name)

        if condition is not None:
            result['condition'] = dict()
            result['condition']['original'], result['condition']['new'] = self.set_select(3, condition)

        if category is not None:
            result['category'] = dict()
            result['category']['original'], result['category']['new'] = self.set_select(2, category)

        if sign is not None:
            result['sign'] = dict()
            result['sign']['original'], result['sign']['new'] = self.set_select(1, sign)

        if area is not None:
            result['area'] = dict()
            result['area']['original'], result['area']['new'] = self.set_select(0, area)

        selector_ = [{"title": "Редактировать запись", "index": 0},
                     {"title": "Ввод", "control_type": "Button", "index": 0}]
        self.find_element(selector_).click_input()

        return result

    def shift_f4(self, set_type, attr_value=None, click_visible=False, click_editable=False, click_required=False):
        selector_ = [{"title": f"\"{self.title}\"", "index": 0}]
        self.find_element(selector_).type_keys(self.SHIFT_EDIT)
        result = {}

        if attr_value is not None:
            result['attr_value'] = dict()
            result['attr_value']['original'], result['attr_value']['new'] = set_type(1, attr_value)

        if click_visible:
            selector_ = [{"title": "Редактировать запись", "index": 0},
                         {"title": "Показывать в диалоге", "control_type": "Pane", "index": 0}]
            self.find_element(selector_).click_input()

        if click_editable:
            selector_ = [{"title": "Редактировать запись", "index": 0},
                         {"title": "Разрешить редактировать?", "control_type": "Pane", "index": 0}]
            self.find_element(selector_).click_input()

        if click_required:
            selector_ = [{"title": "Редактировать запись", "index": 0},
                         {"title": "Вводить обязательно?", "control_type": "Pane", "index": 0}]
            self.find_element(selector_).click_input()

        selector_ = [{"title": "Редактировать запись", "index": 0},
                     {"title": "Ввод", "control_type": "Button", "index": 0}]
        self.find_element(selector_).click_input()

        return result

    # ? FILTER ROW

    def filter(self, index, category, value, left=0):
        selector_ = [{"title": f"\"{self.title}\"", "index": 0},
                     {"title": "", "control_type": "Pane", "index": index}]
        self.find_element(selector_).click_input(coords=(30, 30))
        self.find_element(selector_).type_keys(self.FILTER_RESET)

        self.find_element(selector_).click_input(coords=(30, 30))
        self.find_element(selector_).type_keys(self.Keys.LEFT * 10)
        self.find_element(selector_).type_keys(self.FILTER)

        selector_ = [{"title": "Выборка по запросу", "index": 0},
                     {"title": category, "control_type": "ListItem", "index": 0}]
        self.find_element(selector_).click_input()

        selector_ = [{"title": "Выборка по запросу", "index": 0},
                     {"title": "", "control_type": "Edit", "index": 0}]
        self.find_element(selector_).click_input()
        self.find_element(selector_).type_keys(self.format_str(value), with_spaces=True)

        selector_ = [{"title": "Выборка по запросу", "index": 0},
                     {"title": "Ввод", "control_type": "Button", "index": 0}]
        self.find_element(selector_).click_input()
        self.wait_element(selector_, appear=False)

        selector_ = [{"title": f"\"{self.title}\"", "index": 0},
                     {"title": "", "control_type": "Pane", "index": index}]
        self.find_element(selector_).click_input()
        self.find_element(selector_).type_keys(str(self.Keys.LEFT * left) + self.COPY_VALUE)

        self.clipboard_get(value_error=f'не найдена кат.: {value}')

    # ? SEARCH ROW

    def search(self, index, category, value):
        selector_ = [{"title": f"\"{self.title}\"", "index": 0},
                     {"title": "", "control_type": "Pane", "index": index}]
        self.find_element(selector_).click_input(coords=(30, 30))
        self.find_element(selector_).type_keys(self.FILTER_RESET)

        self.find_element(selector_).click_input(coords=(30, 30))
        self.find_element(selector_).type_keys(self.Keys.LEFT * 10)
        self.find_element(selector_).type_keys(self.SEARCH)
        if not self.wait_element([{"title": "Поиск", "index": 0}]):
            raise Exception('Окно поиска не открылось.')

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "", "control_type": "Edit", "index": 1}]
        self.find_element(selector_).click_input()

        selector_ = [{"title": "", "index": 0},
                     {"title": "", "control_type": "ListItem", "index": category}]
        self.find_element(selector_).click_input()

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "", "control_type": "Edit", "index": 0}]
        self.find_element(selector_).type_keys(self.format_str(value, enter=True), with_spaces=True)

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "", "control_type": "Pane", "index": 4}]
        self.find_element(selector_).type_keys(self.COPY_VALUE)
        self.clipboard_get(value_error=f'не найдена кат.: {value}')

        selector_ = [{"title": "Поиск", "index": 0},
                     {"title": "Перейти", "control_type": "Button", "index": 0}]
        self.find_element(selector_).click_input()
        self.wait_element(selector_, appear=False)

    # ? MOVE BUTTONS

    def move_to_right(self, index=1):
        selector_ = [{"title": f"\"{self.title}\"", "index": 0},
                     {"title": "", "control_type": "Button", "index": index}]
        self.find_element(selector_).click_input()

    def move_to_left(self, index=0):
        selector_ = [{"title": f"\"{self.title}\"", "index": 0},
                     {"title": "", "control_type": "Button", "index": index}]
        self.find_element(selector_).click_input()

        if index == 0:
            selector_ = [{"title": "Новая запись", "index": 0},
                         {"title": "Ввод", "control_type": "Button", "index": 0}]
            self.find_element(selector_).click_input()


class Xls:
    def __init__(self, file_path: str, sheet_name: str = None):
        self.file_path = file_path
        self.wb = xlrd.open_workbook(self.file_path)
        self.ws = self.wb.sheet_by_name(sheet_name) if sheet_name is not None else self.wb.sheet_by_index(0)

    def find(self, value):
        rows = list(self.ws.get_rows())
        pairs = list()
        for n, row in enumerate(rows):
            col_indexes = [(n, i) for i, x in enumerate(row) if x.value == value]
            pairs = [*pairs, *col_indexes]
        return pairs

    def get(self, rowx, colx):
        return self.ws.cell_value(rowx=rowx, colx=colx)


class Excel:
    def __init__(self, file_path: str, sheet_name: str = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        if Path(file_path).is_file():
            self.wb = load_workbook(self.file_path)
        else:
            self.wb = Workbook()
            self.wb.active.title = self.sheet_name
            self.save()
        if self.sheet_name is not None:
            if self.sheet_name in self.wb.sheetnames:
                self.ws = self.wb[self.sheet_name]
            else:
                self.wb.create_sheet(self.sheet_name)
                self.ws = self.wb[self.sheet_name]
                self.save()
        else:
            self.ws = self.wb.active

    def __del__(self):
        self.wb.close()

    def init(self, months):
        rows = [row[1] for row in list(self.ws.values)]
        if len(rows) > 1:
            return self
        for n, month in enumerate(months):
            self.ws.cell(1, n + 4).value = month
        self.save()
        return self

    def add_branch(self, branch, values, index):
        rows = [row[0] for row in list(self.ws.values) if row[0]]
        if index not in rows:
            last_index = self.ws.max_row
            self.ws.cell(last_index + 1, 1).value = int(index)
            self.ws.merge_cells(start_row=last_index + 1, start_column=1, end_row=last_index + 3, end_column=1)
            self.ws.cell(last_index + 1, 2).value = branch
            self.ws.merge_cells(start_row=last_index + 1, start_column=2, end_row=last_index + 3, end_column=2)
            for n, key in enumerate(values.keys()):
                self.ws.cell(last_index + 1 + n, 3).value = key
            self.save()
        return self

    def fill(self, data):
        self.add_branch(data['branch'], data['values'], data['index'])
        row_ = self.find(data['index'], col_index=1)
        col_ = self.find(data['month'], row_index=1)
        row_index = row_[0][0]
        col_index = col_[0][1]
        print(list(data['values'].keys()))
        if data['index'] == 777:
            print()
        self.ws.cell(row_index, col_index).number_format = '# ##0'
        self.ws.cell(row_index, col_index).value = data['values'][list(data['values'].keys())[0]]
        self.ws.cell(row_index + 1, col_index).number_format = '# ##0'
        self.ws.cell(row_index + 1, col_index).value = data['values'][list(data['values'].keys())[1]]
        self.ws.cell(row_index + 2, col_index).number_format = '# ##0'
        self.ws.cell(row_index + 2, col_index).value = data['values'][list(data['values'].keys())[2]]
        self.save()

    def find(self, value, row_index=None, col_index=None):
        rows = [row for row in list(self.ws.values)]
        pairs = list()
        for n, row in enumerate(rows):
            col_indexes = [(n + 1, i + 1) for i, x in enumerate(row) if x == value]
            pairs = [*pairs, *col_indexes]
        if row_index:
            pairs = [pair for pair in pairs if pair[0] == row_index]
        if col_index:
            pairs = [pair for pair in pairs if pair[1] == col_index]
        return pairs

    def set(self, row, col, value):
        self.ws.cell(row=row, column=col).value = value

    def save(self):
        if Path(self.file_path).is_file():
            shutil.copy(self.file_path.__str__(), self.file_path.__str__().replace('.xlsx', '_b.xlsx'))
        return self.wb.save(self.file_path)

    def close(self):
        return self.wb.close()


if __name__ == '__main__':
    pass
